#!/usr/bin/env python3
# -*- coding: utf-8 -*-
file_name="Store Device Connections Checker"
version="1.0.1"
file_details=f"""
##################################################################################################################
                            __filename__ = "{file_name}_{version}.py"
                            __author__ = "Developed by Jayakar,Kathiravan,Rohan"
                            __copyright__ = "Copyright (C) 2025, H&M Store Network Team"
                            __credits__ = ["Kathiravan A C", "Jayakar Indrakumar", "Rohan Sakthivel"]
                            __license__ = "H&M group"
                            __version__ = "{version}"
                            __date__ = "15-03-2025"
                            __maintainer__ = "KJR"
                            __status__ = "Working"
##################################################################################################################
 
"""
print(file_details)
 
 
from netmiko import ConnectHandler
import paramiko
import re
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd
import os
from openpyxl import load_workbook
from getpass import getpass
import requests
import urllib3
from rich.console import Console
from xlsxwriter import Workbook
import json
from openpyxl.styles import PatternFill

#########################################
india_timezone = ZoneInfo("Asia/Kolkata")
urllib3.disable_warnings()
#########################################
 
class store:
    store_access_token=""
    base_url="https://apigw-eucentral2.central.arubanetworks.com"
    def getToken():
        return store.store_access_token
   
def refreshToken():
    global access_token
    access_token=input("Enter the access token: ")
    store.store_access_token=access_token
    print("Access token updated successfully. 🔃")
#########################################


class reqs:
    def __init__(self,url=None,data=None):
        self.url=f"{store.base_url}{url}"
        self.data=data
        self.total_count=0
        self.offset=0
   
    @property
    def apiHeaders(self):
        return {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {store.getToken()}"
        }
 
    def get(self):
        while True:
            try:
                req = requests.get(self.url,headers=self.apiHeaders, verify=False)
                response= req.json()
                return response[self.data]
            except Exception as e:
                if response.get("error")=='invalid_token':
                    refreshToken()
                elif str(e)==f"'{self.data}'":
                    print(f"Invalid data : {e} ")
                print(f"{e} : {response}\nRetrying...🔄")
   
    def getdf(self):
        while True:
            try:
                req = requests.get(self.url,headers=self.apiHeaders, verify=False)
                response= req.json()
                if "total" in self.url:
                    self.total_count = int(response["total"])
                self.df = pd.DataFrame.from_dict(response[self.data])
                return self.df
            except Exception as e:
                if response.get("error")=='invalid_token':
                    refreshToken()
                elif str(e)==f"'{self.data}'":
                    print(f"Invalid data : {e} ")
                print(f"{e} : {response}\nRetrying...🔄")
 
    def getall(self):
        self.getdf()
        df_master = None
        print(self.total_count)
        for obj in range(int(self.total_count/1000)+1):
            self.offset+=1000
            if "offset" in self.url:
                self.url=re.sub(r'offset=\d+', f'offset={self.offset}', self.url)
            if df_master is None:
                df_master = self.df
            else:
                df_master = pd.concat([df_master,self.df])
            self.getdf()
        return df_master
    
    def text(self):
        while True:
            try:
                req = requests.get(self.url,headers=self.apiHeaders, verify=False)
                response= req.text  
               
                if req.status_code != 200:
                    raise Exception("Request failed... ❌")
                return response
            except Exception as e:
                response=req.json()
                if response.get("error")=='invalid_token':
                    refreshToken()
                elif str(e)==f"'{self.data}'":
                    print(f"Invalid data : {e} 🔄")
 
#####################################
class SSHoutput:
    def __init__(self,devices,commands,device_type):
        self.devices=devices
        self.commands=commands
        self.devices_type=device_type
   
    def output_txt(self,file_path):
        for device_ip in self.devices:
            self.username=jump_username
            self.password=jump_password
            for _ in range(3):
                try:
                    vmtransport = vm.get_transport()
                    dest_addr = (device_ip, 22)
                    local_addr = (jump_ip, 22)
                    vmchannel = vmtransport.open_channel("direct-tcpip", dest_addr, local_addr)
                    ##############################
                    device = {
                                'device_type': 'aruba_os',
                                'host': device_ip,
                                'username': self.username,
                                'password': self.password,
                                'secret': jump_password,
                                'port': 22,
                                'verbose': True,
                                'sock': vmchannel
                            }
                    ##############################
                    connection = ConnectHandler(**device)
                    connection.enable()
                    output = connection.send_multiline_timing(self.commands)
                    ##############################
                    print(output)
                    ##############################
                except Exception as e:
                    print(f"Failed to connect to {device_ip}: {e}")
                    print("Retrying with stradmin...")
                    if self.devices_type=="gateway":
                        self.username="stradmin"
                        self.password="@ru8@N1(hp#"
                    elif self.devices_type=="switch":
                        self.username="stradmin"
                        self.password="H@V#@n!(e#@y"
                    pass
                else:
                    break
               
            ##############################
            try:
                with open(file_path, "a") as file:
                    log_entry = f"{output}\n"
                    file.write(log_entry)
 
            except Exception as e:
                print(f"Failed to save the output of {device_ip}: {e}")
                pass
            ##############################
#########################################



class backup:
    def pre_log():
        wired_url = f"/monitoring/v1/clients/wired?site={site_name}"
        wired_df=reqs(url=wired_url,data="clients").getdf()
        wireless_url = f"/monitoring/v1/clients/wireless?site={site_name}"
        wireless_df=reqs(url=wireless_url,data="clients").getdf()
        pre_df=pd.concat([wired_df,wireless_df],ignore_index=True)
        ap_url= f"/monitoring/v2/aps?site={site_name}&status=Up"
        ap_df= reqs(ap_url,"aps").getdf()
        ap_df_filtered = ap_df.rename(columns = {"serial":"associated_device","macaddr":"associated_device_mac","name":"associated_device_name"})
        ap_df_filtered = ap_df_filtered[["associated_device","associated_device_mac","associated_device_name","ip_address"]]
        pre_df=pd.concat([pre_df,ap_df_filtered],ignore_index=True)

        user_role_counts = pre_df["user_role"].value_counts().reset_index()
        user_role_counts.columns = ["user_role", "Count_Before"]


        folder_name = f"Site_logs"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        excel_file_path = os.path.join(folder_name,f"{site_name}.xlsx")
        with pd.ExcelWriter(excel_file_path) as writer:
            pre_df.to_excel(writer, sheet_name="Clients Data", index=False)
            user_role_counts.to_excel(writer, sheet_name="User Role Counts", index=False)
        print(f"The post-validation excel for {site_name} was successfully saved in {folder_name}. 📄✅")

        ##############################
        folder_name_log = f"{site_name}_Backup"
        log_file_path = os.path.join(folder_name_log, f"{site_name}_Backup_Logs_{india_time}.txt")
        if not os.path.exists(log_file_path):
            with open(log_file_path, "w") as file:
                file.write(f"Pre-Logs-{india_time}")
        print(f"The backup-log file for {site_name} was successfully created 📄, but the log data has not been saved yet. ⏸️")

        '''
        if not os.path.exists(excel_file_path):
            with open(excel_file_path, "w") as file:
                file.write(f"Pre-Logs-{india_time}")
        print(f"The backup-log file for {site_name} was successfully created 📄, but the log data has not been saved yet. ⏸️")'''
        ##############################
 
        gateway_log=SSHoutput(gateway_ip,gateway_commands,device_type="gateway")
        gateway_log.output_txt(file_path=log_file_path)
        print(f"The {site_name} gateways' backup log has been uploaded successfully. 📤")
  
        switch_log=SSHoutput(switch_ip,switch_commands,device_type="switch")
        switch_log.output_txt(file_path=log_file_path)
        print(f"The {site_name} switches' backup log has been uploaded successfully. 📤")
        ##############################

    def config_status(switch_serial):
        config_details = f"/configuration/v1/devices/{switch_serial}/config_details"
        response = reqs(url=config_details).text()
        start = response.find('{')  
        end = response.rfind('}')  
        json_part = response[start:end+1] 
        #print(json_part)
        json_json = json.loads(json_part)
        #print(json_json)
        json_df = pd.DataFrame([json_json])
        json_config = json_df.at[0,"Configuration_error_status"]
        #print(json_config)
        return json_config
    
class switch_name:
    def __init__(self):
        pass
    def name_switch(self,switch_port_serial):
        self.switch_port_serial = switch_port_serial
        #print(self.switch_port_serial)
        swi_name_dict={}
        for serial in self.switch_port_serial:
            test_url=f"/monitoring/v1/switches/{serial}"
            swi_name=reqs(test_url,"name").get()
           
            swi_name_dict[serial] = swi_name
        return swi_name_dict      
     
class switch_type:
    def __init__(self):
        pass
    def no_ports(self,switch_port_serial):
        self.switch_port_serial = switch_port_serial
        #print(self.switch_port_serial)
        l1=[]
        for serial in self.switch_port_serial:
            test_url=f"/monitoring/v1/switches/{serial}/ports"
            bounce_port=reqs(test_url,"count").get()
            l1.append(bounce_port)
        return l1
            #validation().port_bounce_all_ports(bounce_port,serial)


def highlight_sheet_tab(worksheet, workbook, color="FF0000"):
    """
    Highlight the tab of a worksheet with the specified color.
    Args:
        worksheet: The worksheet object
        workbook: The workbook object
        color: The color code (default is red FF0000)
    """
    sheet_name = worksheet.title
    sheet_index = None
    for idx, sheet in enumerate(workbook.worksheets):
        if sheet.title == sheet_name:
            sheet_index = idx
            break
    
    if sheet_index is not None:
        workbook.worksheets[sheet_index].sheet_properties.tabColor = color
                    
def should_highlight(port, switch_type):
    try:
        port = int(port)
    except (TypeError, ValueError):
        return True  
    
    if switch_type == "Gateway" and port in [0, 1]:
        return False
    elif switch_type == "24 port switch" and 23 <= port <= 28:
        return False
    elif switch_type == "48 port switch" and 46 <= port <= 52:
        return False
    else:
        return True

def should_highlight_down(value):
    if value != 0:
        return True
    else:
        return False

def should_highlight_sync(value):
    if value == "In Sync":
        return False
    else:
        return True
       

def device_status(site_name):

    gateway_url = f"/monitoring/v1/gateways?site={site_name}"
    gateway_df=reqs(gateway_url,"gateways").getdf()
    #gateway_df['status_message'] = gateway_df.apply(check_status, axis=1)
    gateway_status_df = gateway_df[["name","status"]]
    gateway_down_count = gateway_status_df[gateway_status_df["status"] == "Down"].shape[0]
    gateway_down_count_df = pd.DataFrame({'gateway down': [gateway_down_count]})
    
    
    switch_url= f"/monitoring/v1/switches?site={site_name}"
    switch_df= reqs(switch_url,"switches").getdf()
    switch_serial = list(switch_df["serial"])
    #switch_df['status_message'] = switch_df.apply(check_status, axis=1)
    '''sync_status_list = []
    for i in switch_df["serial"]:
        switch_config_url = f"/configuration/v1/devices/{i}/config_details"
        switch_config_str = reqs(switch_config_url,"Configuration_error_status").get()
        #sync_status_list.append(switch_config_str) 
    switch_df["sync status"] = sync_status_list
    '''
    switch_status_df = switch_df[["name","status"]]
    switch_down_count = switch_status_df[switch_status_df["status"] == "Down"].shape[0]
    switch_down_count_df = pd.DataFrame({'switch down': [switch_down_count]})
    config_status_list = []
    switch_name_dict =switch_name().name_switch(switch_serial)
    for j in switch_serial:
        config_stat =  backup.config_status(j)    
        status = "Not in Sync" if config_stat else "In Sync"
        config_status_list.append({"switch_name": switch_name_dict[j], "config_status": status})
    config_stat_df = pd.DataFrame(config_status_list)
    print(switch_status_df)
    
    
    ap_url= f"/monitoring/v2/aps?site={site_name}"
    ap_df= reqs(ap_url,"aps").getdf()
    #ap_df['status_message'] = ap_df.apply(check_status, axis=1)
    ap_status_df = ap_df[["name","status"]]
    ap_down_count = ap_status_df[ap_status_df["status"] == "Down"].shape[0]
    ap_down_count_df = pd.DataFrame({'ap down': [ap_down_count]})
    print(ap_status_df)
    combined_df = pd.concat([gateway_down_count_df,switch_down_count_df,ap_down_count_df,config_stat_df],axis = 1)
    
    return combined_df
#######################################

while True:
    try:
        jump_ip = input("Enter your Jump Server ip address: ")
        jump_username = input("Enter your Jump Server username: ")
        jump_password = getpass("Enter your Jump Server Password:")
        vm = paramiko.SSHClient()
        vm.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        vm.connect(jump_ip, username=jump_username, password=jump_password, timeout=10)
    except Exception as e:
        print(f"Failed to connect to {jump_ip} ❌: {e}\nRetry...🔄")
    else:
        break
#########################################

site_name = list(map(str,input("Enter the site name: ").split(",")))
refreshToken()

def store_device_connection(site_name):
    os.makedirs(f"Site_logs", exist_ok=True)
    india_time = datetime.now(india_timezone).strftime('%H%M%S_%d%m%Y')
    name_excel = f"{site_name[0]} - {site_name[len(site_name)-1]}"
    excel_filename = os.path.join(f'Site_logs', f'Site_Details_{name_excel}.xlsx')
    #file_path = f"{site_name}_Backup_Logs_{india_time}.txt"
    #excel_filename = f"Network_Connections_{india_time}.xlsx"
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
    good_connections=[]
    bad_connections=[]
    for i in site_name:
        site_name=i.strip()
        #file_path = f"{i}_Backup_Logs_{india_time}.txt"
        while True:
            try:
                gateway_url = f"/monitoring/v1/gateways?site={site_name}"
                gateway_df=reqs(gateway_url,"gateways").getdf()

                if "ip_address" in gateway_df.columns:
                    gateway_ip = list(gateway_df["ip_address"])
                    gateway_serial=list(gateway_df["serial"])
                    
                    gateway_type = "Gateway"  # Default 
                    print(gateway_serial)
                else:
                    gateway_ip = []
                    gateway_serial=[]
                    gateway_name = []
                ############################################################
                switch_url= f"/monitoring/v1/switches?site={site_name}"
                switch_df= reqs(switch_url,"switches").getdf()
                if "ip_address" in switch_df.columns:
                    switch_ip = list(switch_df["ip_address"])
                    switch_serial=list(switch_df["serial"])
                    print(switch_serial)
                    switchType = switch_type().no_ports(switch_serial)
                    switch_name_dict =switch_name().name_switch(switch_serial)
                    switch_type_dict = {}
                    for idx, serial in enumerate(switch_serial):
                        if idx < len(switchType):
                            current_type = str(switchType[idx]).strip()
                            print(f"Switch {serial} has type: '{current_type}'")
                            
                            if current_type == "52" or current_type =="48" or "52" in current_type:
                                switch_type_dict[serial] = "48 port switch"
                            elif current_type == "28"  or "28" in current_type:
                                switch_type_dict[serial] = "24 port switch"
                            else:
                                switch_type_dict[serial] = "Unknown"
                    print("Switch type dictionary:", switch_type_dict)
                else:
                    switch_ip=[]
                    switch_serial=[]
                    switch_type_dict = {}
                ############################################################
                all_connections = []
                l1=[]
            
                for switch in switch_serial:
                    gate_switch_port_url = f"/topology_external_api/edges/{switch}/{gateway_serial[0]}"
                    gate_switch_port_df= reqs(gate_switch_port_url,"edges").get()
                    
                    if gate_switch_port_df:
                        try:
                            edges_list = gate_switch_port_df["other"]["edges"]
                            print('Gateway to switch port connections:')
                            for edge in edges_list:
                                from_port = edge["fromIf"]["portNumber"]
                                to_port = edge["toIf"]["portNumber"]
                                all_connections.append({
                                    "Device 1 Name": f"{site_name.lower()}vpr001",
                                    "Device 1 Type": gateway_type,
                                    "From Port no": from_port,
                                    "Device 2 Name": switch_name_dict[switch],
                                    "Device 2 Type": switch_type_dict.get(switch, "Unknown"),
                                    "To Port no": to_port,
                                    "Connection-type": "Gateway to switch"                                        
                                })
                                
                        except KeyError as e:
                            print(f"KeyError: {e}. The expected structure is missing in the API response.")
                    else:
                        l1.append(switch)
                        
                print(l1)
                ############################################################
                l3=switch_serial
                l2=[]      #23 24 25 26
                l2=l3      #46 47 48
                l4=[]
                for switch1 in l3:
                    for switch2 in l2:
                        if switch1 != switch2:
                            if [switch1,switch2] not in l4 and [switch2,switch1] not in l4:
                                l4.append([switch1,switch2])
                                gate_switch_port_url_1 = f"/topology_external_api/edges/{switch1}/{switch2}"
                                gate_switch_port_df_1= reqs(gate_switch_port_url_1,"edges").get()
                                if gate_switch_port_df_1:
                                    try:
                                        edges_list1 = gate_switch_port_df_1["other"]["edges"]
                                        print("switch to switch port connections:")
                                        for edge1 in edges_list1:
                                            from_port1 = edge1["fromIf"]["portNumber"]
                                            to_port1 = edge1["toIf"]["portNumber"]
                                            all_connections.append({
                                                "Device 1 Name": switch_name_dict[switch1],
                                                "Device 1 Type": switch_type_dict.get(switch1, "Unknown"),
                                                "From Port no": from_port1,
                                                "Device 2 Name": switch_name_dict[switch2],
                                                "Device 2 Type": switch_type_dict.get(switch2, "Unknown"),
                                                "To Port no": to_port1,
                                                "Connection-type": "Switch to switch"                                        
                                            })
                                            
                                    except KeyError as e:
                                        print(f"KeyError: {e}. The expected structure is missing in the API response.")
                                else:
                                    pass
                ############################################################                
                result_df1 = pd.DataFrame(all_connections)
                result_df2 = device_status(i)

                df1_reset = result_df1.reset_index(drop=True)
                df2_reset = result_df2.reset_index(drop=True)
                seperator = pd.DataFrame([['', '']])
                result_new = pd.concat([df1_reset,df2_reset],axis =1)

                result_new.to_excel(writer, sheet_name=f"{site_name}", index=False)
                #result_excel.to_excel(writer, sheet_name=f"{site_name}", index=False)

                workbook = writer.book
                worksheet = writer.sheets[site_name]

                has_highlighted_cells = False
                try:
                    """wb = load_workbook("Result.xlsx")
                    ws = wb[f'{site_name}']"""
                    
                    highlight_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    
                    headers = list(result_new.columns)
                    source_port_col = headers.index("From Port no") + 1  
                    dest_port_col = headers.index("To Port no") + 1
                    switch1_type_col = headers.index("Device 1 Type") + 1
                    switch2_type_col = headers.index("Device 2 Type") + 1
                    gateway_down_col = headers.index("gateway down") + 1
                    switch_down_col = headers.index("switch down") + 1
                    ap_down_col = headers.index("ap down") + 1
                    config_stat_col = headers.index("config_status") + 1
                    
                    for row in range(2, len(result_new) + 2):  
                        switch1_type = worksheet.cell(row=row, column=switch1_type_col).value
                        switch2_type = worksheet.cell(row=row, column=switch2_type_col).value
                        source_port = worksheet.cell(row=row, column=source_port_col).value
                        dest_port = worksheet.cell(row=row, column=dest_port_col).value
                        
                        
                        print(f"Row {row}: {switch1_type} port {source_port}, {switch2_type} port {dest_port}")
                        
                        source_port_cell = worksheet.cell(row=row, column=source_port_col)
                        if should_highlight(source_port, switch1_type):
                            source_port_cell.fill = highlight_fill
                            print(f"  Highlighting source port {source_port}")
                            has_highlighted_cells = True
                        
                        dest_port_cell = worksheet.cell(row=row, column=dest_port_col)
                        if should_highlight(dest_port, switch2_type):
                            dest_port_cell.fill = highlight_fill
                            print(f"  Highlighting destination port {dest_port}")
                            has_highlighted_cells = True

                    for row in range(2, 3):     
                        gateway_down = worksheet.cell(row=row,column=gateway_down_col).value
                        switch_down = worksheet.cell(row=row,column=switch_down_col).value
                        ap_down = worksheet.cell(row=row,column=ap_down_col).value
                        
                        gateway_down_cell = worksheet.cell(row=row, column=gateway_down_col)
                        if should_highlight_down(gateway_down):
                            gateway_down_cell.fill = highlight_fill
                            print(f"  Highlighting gateway {gateway_down}")
                            has_highlighted_cells = True

                        switch_down_cell = worksheet.cell(row=row, column=switch_down_col)
                        if should_highlight_down(switch_down):
                            switch_down_cell.fill = highlight_fill
                            print(f"  Highlighting switch {switch_down}")
                            has_highlighted_cells = True
                        
                        ap_down_cell = worksheet.cell(row=row, column=ap_down_col)
                        if should_highlight_down(ap_down):
                            ap_down_cell.fill = highlight_fill
                            print(f"  Highlighting ap {ap_down}")
                            has_highlighted_cells = True
                    
                    for row in range(2,len(result_new["config_status"])+1):
                        config_stat = worksheet.cell(row=row,column=config_stat_col).value
                        
                        config_stat_cell = worksheet.cell(row=row, column=config_stat_col)
                        if should_highlight_sync(config_stat):
                            config_stat_cell.fill = highlight_fill
                            print(f"  Highlighting switch {config_stat}")
                            has_highlighted_cells = True


                    if has_highlighted_cells:
                        highlight_sheet_tab(worksheet, workbook, "FF0000")
                        bad_connections.append(site_name)
                    else:
                        good_connections.append(site_name)
                    
                    # Save the changes
                    #wb.save("Result.xlsx")
                    print(f"Excel file saved with highlighted ports at {excel_filename}")
                except Exception as e:
                    print(f"Error applying conditional formatting: {e}")
                    import traceback
                    traceback.print_exc()
                    print(f"Excel file saved without highlighting at {excel_filename}")
                break
        
            ##############################
            except Exception as e:
                print(f"Invalid access token: {e}\nRetry...")
                if True:
                    access_token = input("Enter access token: ")
                else:
                    break

        ##############################
    writer.close()
    print(f"All sites saved to {excel_filename}")
    os.makedirs(f"Site_logs", exist_ok=True)
    india_time = datetime.now(india_timezone).strftime('%H%M%S_%m%d%Y')
    
    if good_connections != []:
        good_connections_df= pd.DataFrame(good_connections)
        good_connections_df.columns = ["Good sites"]
        os.makedirs(f"Site_logs/Good_sites",exist_ok =True)
        excel_good_filename = os.path.join(f'Site_logs/Good_sites',"Good sites for migration.xlsx")
        good_connections_df.to_excel(excel_good_filename,index=False)
        

    if bad_connections != []:

        os.makedirs(f"Site_logs/Bad_sites",exist_ok =True)
        bad_connections_df= pd.DataFrame(bad_connections)
        bad_connections_df.columns = ["Bad sites"]
        excel_bad_filename = os.path.join(f'Site_logs/Bad_sites',"Bad sites.xlsx")
        bad_connections_df.to_excel(excel_bad_filename,index=False)
    
    return good_connections
    ##############################
 
india_time = datetime.now(india_timezone).strftime('%H%M%S_%m%d%Y')

store_device_connection(site_name)

gateway_commands=["sh lldp neigh",
"show user-table",
"show station-table",
"show ip interface brief",
"show ip oap advertise",
"show arp",
"show port status",
"show uplink debug",
"show ip oap tunnel",
"show user",
"show boot history",
"show image version",
"show switches",
"show running-config",
"show version",
"show clock"
]
 
switch_commands=["show port-access clients",
"show cdp neighbors",
"show interface status",
"show interface transceiver det",
"show version",
"show flash",
"show running-config"
]
 
#good_sites = store_device_connection(site_name)



for i in site_name:
    site_name=i.strip()

    gateway_url = f"/monitoring/v1/gateways?site={site_name}"
    gateway_df=reqs(gateway_url,"gateways").getdf()
    gateway_ip = list(gateway_df["ip_address"])
    gateway_serial=list(gateway_df["serial"])
    print(gateway_ip)

    switch_url= f"/monitoring/v1/switches?&site={site_name}"
    switch_df= reqs(switch_url,"switches").getdf()
    config_status_list = []
    switch_df_up = switch_df[switch_df["status"]=="Up"]
    switch_name_1 =list(switch_df_up["name"])
    switch_ip = list(switch_df_up["ip_address"])
    switch_serial  =list(switch_df_up["serial"])
    print(switch_ip)
    
    for j in switch_serial:
        config_stat =  backup.config_status(j)
        status = "Not in Sync" if config_stat else "In Sync"
        '''s = pd.Series([config_stat], index=["Configuration_error_status"])
        value = s.iloc[0]'''
        config_status_list.append({"serial": j, "config_status": status})
    config_stat_df = pd.DataFrame(config_status_list)
    switch_df_merge = pd.merge(switch_df,config_stat_df,how="left",on="serial")
    ap_url= f"/monitoring/v2/aps?site={site_name}"
    ap_df= reqs(ap_url,"aps").getdf()
        
    
    folder_name = f"{site_name}_Backup"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    excel_file = os.path.join(folder_name, f"{site_name}_Details.xlsx")
    with pd.ExcelWriter(excel_file, engine="xlsxwriter") as writer:
        gateway_df.to_excel(writer, sheet_name="Gateway Details", index=False)
        switch_df_merge.to_excel(writer, sheet_name="Switch Details", index=False)
        ap_df.to_excel(writer, sheet_name="AP Details", index=False)


    backup.pre_log()
print(f"Retrieving the {site_name} backup log has been successfully completed. ✅")
print(f"Device status for the site {site_name} has been uploaded successfully. ✅")


